import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor
from openpyxl.utils import get_column_letter
import winsound

# web scraping functions
# driver is started locally (and ended)
# get_html_from_url() uses sleep=10 with puuilo.fi and tokmanni.fi as per their robots.txt
class Scraper:
    def __init__(self):
        self.driver = self.local_driver()

    def local_driver(self):
        chrome_options = Options()
        chrome_options.add_argument('--headless')  # Run Chrome in headless mode (without UI)
        service = ChromeService(executable_path=ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        return driver

    def end_session(self):
        self.driver.quit()

    def get_html_from_url(self, url, sleep=1):
        try:
            self.driver.get(url)
            time.sleep(sleep)
            return self.driver.page_source
        except Exception as e:
            print(f"Error getting HTML from {url}: {str(e)}")
            return None

class PrismaScraper(Scraper):
    def __init__(self):
        super().__init__()
    def search_product(self, product_code):
        try:
            url = f"https://www.prisma.fi/haku?search={product_code}"
            html = self.get_html_from_url(url)
            if html:
                soup = BeautifulSoup(html, 'html.parser')
                if not self.search_result_check(soup):
                    #print("\nPRISMA:\nEi tuloksia haulla:", product_code)
                    return None, None, None
                link = self.specific_product_page(soup)
                link = f"https://www.prisma.fi{link}"
                html = self.get_html_from_url(link)
                if html:
                    soup = BeautifulSoup(html, 'html.parser')
                    product_name, brand_name, price = self.extract_product_information(soup)
                    if product_name and price:
                        #print("\nPRISMA:\ntuotekoodi:", product_code,"\ntuotenimi:", product_name,"\nhinta:", price, "€")
                        return product_name, link, brand_name
                else:
                    print("(1) PRISMA: Error with handling HTML.")
                    return None, None, None
            else:
                print("(2) PRISMA: Error with handling HTML.")
                return None, None, None
        except Exception as e:
            return None, None, None

    def search_result_check(self, soup): #returns False if there were no search results or True if there were results
        try:
            no_search_results_element = soup.find('p', {'data-test-id': 'no-search-results', 'class': 'search_noResultTitle__PTo4Q'})
            if no_search_results_element:
                result_text = no_search_results_element.get_text(strip=True)
                return False
            else:
                return True
        except Exception as e:
            return None

    def specific_product_page(self, soup):
        try:
            link_tag = soup.find('a', class_='ProductCard_blockLink__NTUGB') # Find the <a> tag with the specified class
            if link_tag:
                link = link_tag['href'] # Extract the value of the href attribute
                return link
            else:
                return None
        except Exception as e:
            return None
        
    def extract_product_information(self, soup):
        try:
            # Find the element with class 'ProductMainInfo_brand__NbMHp'
            brand_div = soup.find('div', class_='ProductMainInfo_brand__NbMHp')
            # Extract the brand text from the <a> tag inside the div
            brand_name = brand_div.find('a').text

            # Find the element with class 'ProductMainInfo_finalPrice__1hFhA'
            price_span = soup.find('span', class_='ProductMainInfo_finalPrice__1hFhA')
            # Extract the price text from the span and remove non-numeric characters
            price_text = price_span.text.replace('&nbsp;', '').replace('€', '').replace(',', '.')
            # Convert the extracted text to a float
            price = float(price_text)

            # Find the element with class 'ProductMainInfo_title__VKU68' and data-test-id 'product-name'
            product_name_h1 = soup.find('h1', class_='ProductMainInfo_title__VKU68', attrs={'data-test-id': 'product-name'})
            # Extract the text content from the h1 element
            product_name = product_name_h1.text
            
            return product_name, brand_name, price
        except Exception as e:
            return None, None, None

class PuuiloScraper(Scraper):
    def __init__(self):
        super().__init__()
    def search_product(self, product_code):
        try:
            url = f"https://www.puuilo.fi/catalogsearch/result/?q={product_code}"
            html = self.get_html_from_url(url)
            if html:
                soup = BeautifulSoup(html, 'html.parser')
                if not self.search_result_check(soup):
                    #print("\nPUUILO:\nEi tuloksia haulla:", product_code)
                    return None, None, None
                
                link = self.specific_product_page(soup)
                html = self.get_html_from_url(link)
                if html:
                    soup = BeautifulSoup(html, 'html.parser')
                    price, product_name = self.extract_product_information(soup)
                    if product_name and price:
                        #print("\nPUUILO:\ntuotekoodi:", product_code, "\ntuotenimi:", product_name, "\nhinta:", price, "€")
                        return link, price, product_name
                else:
                    print("(1) PUUILO: Error with handling HTML.")
                    return None, None, None
            else:
                print("(2) PUUILO: Error with handling HTML.")
                return None, None, None
        except Exception as e:
            return None, None, None
    
    def search_result_check(self, soup): # returns True if there are search results or False if no results
        count_span = soup.find('span', class_='amsearch-results-count')
        if count_span and count_span.get_text(strip=True) == '(0)':
            return False
        else:
            return True
        
    def specific_product_page(self, soup): # returns link to the prodcut page or None
        try:
            a_element = soup.select_one('article.product-item-info a.product-item-photo') # Find the <a> element inside the <article> tag
            if a_element:
                link = a_element['href'] # Extract the value of the "href" attribute
                return link
            else:
                return None
        except Exception as e:
            return None

    def extract_product_information(self, soup): # returns price and product name or None
        try:
            # Find the element with class 'price' inside the span with id 'product-price-107111'
            price_span = soup.find('span', class_='price')
            # Extract the price text from the span and remove non-numeric characters
            price_text = price_span.text.replace('&nbsp;', '').replace('€', '').replace(',', '.')
            # Convert the extracted text to a float
            price = float(price_text)

            # Extract product name
            product_name = soup.find('span', {'data-ui-id': 'page-title-wrapper'}).text.strip()

            return price, product_name
        except Exception as e:
            return None, None


class KRautaScraper(Scraper):
    def __init__(self):
        super().__init__()
    def search_product(self, product_code):
        try:
            url = f"https://www.k-rauta.fi/etsi?query={product_code}"
            html = self.get_html_from_url(url)
            if html:
                soup = BeautifulSoup(html, 'html.parser')
                if not self.search_result_check(soup):
                    #print("\nK-RAUTA:\nEi tuloksia haulla:", product_code)
                    return None, None, None
                
                link = self.specific_product_page(soup)
                link = f"https://www.k-rauta.fi{link}"
                html = self.get_html_from_url(link)
                if html:
                    soup = BeautifulSoup(html, 'html.parser')
                    price, product_name = self.extract_product_information(soup)
                    if product_name and price:
                        #print("\nK-RAUTA:\ntuotekoodi:", product_code, "\ntuotenimi:", product_name, "\nhinta:", price, "€")
                        return link, price, product_name
                else:
                    print("(1) K-RAUTA: Error with handling HTML.")
                    return None, None, None
            else:
                print("(2) K-RAUTA: Error with handling HTML.")
                return None, None, None
        except Exception as e:
            return None, None, None

    def search_result_check(self, soup): # returns True if there are search results or False if no results
        try:
            no_search_results_element = soup.find('div', class_="empty-search-result")
            if no_search_results_element:
                result_text = no_search_results_element.get_text(strip=True)
                return False
            else:
                return True
        except Exception as e:
            return None
        
    def specific_product_page(self, soup): # returns specific product page or None
        try:
            a_element = soup.find('a', class_='product-card') # Find the <a> element with class "product-card"
            if a_element:
                link = a_element['href'] # Extract the value of the "href" attribute
                return link
            else:
                return None
        except Exception as e:
            return None

    def extract_product_information(self, soup): # returns price and product name or None
        try:
            # Extract price span
            price_span = soup.find('span', class_='price-view-fi__sale-price--prefix')
            price_text = price_span.next_sibling.strip()
            price= float(price_text.replace("€", "").replace(",", "."))

            # Extract the product name
            product_name = soup.find('h1', class_='product-heading__product-name').text.strip()
            return price, product_name
        except Exception as e:
            return None, None

class TokmanniScraper(Scraper):
    def __init__(self):
        super().__init__()
    def search_product(self, product_code):
        try:
            query = f"\"tokmanni\" \"{product_code}\""
            url = f"https://www.google.com/search?q={query}"
            html = self.get_html_from_url(url, 10)
            if html:
                soup = BeautifulSoup(html, 'html.parser')
                if not self.search_result_check(soup):
                    #print("\nTOKMANNI:\nEi tuloksia haulla:", product_code)
                    return None, None, None
                link = self.specific_product_page(soup, product_code)
                if not link:
                    return None, None, None
                html = self.get_html_from_url(link, 10)
                if html:
                    soup = BeautifulSoup(html, 'html.parser')
                    price, product_name = self.extract_product_information(soup)
                    if product_name and price:
                        #print("\nTOKMANNI:\ntuotekoodi:", product_code, "\ntuotenimi:", product_name, "\nhinta:", price, "€")
                        return link, price, product_name
                else:
                    print("(1) TOKMANNI: Error with handling HTML.")
                    return None, None, None
            else:
                print("(2) TOKMANNI: Error with handling HTML.")
                return None, None, None
        except Exception as e:
            return None, None, None

    def search_result_check(self, soup): # returns False if no results found, True if results were found, otherwise None
        try:
            # Define the target text
            target_text = "About 0 results"

            # Find the div element with class "LHJvCe" and id "result-stats"
            result_stats_div = soup.find('div', {'class': 'LHJvCe', 'id': 'result-stats'})

            # Check if the target text is present in the div
            if result_stats_div and target_text in result_stats_div.get_text():
                return False
            else:
                return True
        except Exception as e:
            return None

    def specific_product_page(self, soup, product_code): # Returns the link to the product page ro None
        try:
            # Find the <em> tag
            em_tag = soup.find('em')
            # Check if the text inside the <em> tag is equal to product_code
            if em_tag and int(em_tag.text.strip()) == int(product_code):
                # Find the <a> tag and get the href value
                a_tag = soup.find('a', {'jsname': 'UWckNb'})
                if a_tag:
                    link = a_tag['href']
                    return link
        except Exception as e:
            return None
        
    def extract_product_information(self, soup): # Returns price and product name or None
        try:
            # Find the span element with class "price-wrapper"
            price_wrapper_span = soup.find('span', class_='price-wrapper')
            # Extract the value from the data-price-amount attribute
            price = price_wrapper_span.get('data-price-amount')
            
            # Find the h1 element with class "page-title"
            page_title_h1 = soup.find('h1', {'class': 'page-title'})
            # Extract product name
            product_name = page_title_h1.get_text(strip=True)
            return price, product_name
        except Exception as e:
            return None, None

class BauhausScraper(Scraper):
    def __init__(self):
        super().__init__()
    def search_product(self, product_code):
        try:
            query = f"\"bauhaus\" \"{product_code}\""
            url = f"https://www.google.com/search?q={query}"
            html = self.get_html_from_url(url, 10)
            if html:
                soup = BeautifulSoup(html, 'html.parser')
                if not self.search_result_check(soup):
                    #print("\nBAUHAUS:\nEi tuloksia haulla:", product_code)
                    return None, None, None
                link = self.specific_product_page(soup, product_code)
                if not link:
                    return None, None, None
                html = self.get_html_from_url(link, 10)
                if html:
                    soup = BeautifulSoup(html, 'html.parser')
                    price, product_name = self.extract_product_information(soup)
                    if product_name and price:
                        #print("\nBAUHAUS:\ntuotekoodi:", product_code, "\ntuotenimi:", product_name, "\nhinta:", price, "€")
                        return link, price, product_name
                else:
                    print("(1) BAUHAUS: Error with handling HTML.")
                    return None, None, None
            else:
                print("(2) BAUHAUS: Error with handling HTML.")
                return None, None, None
        except Exception as e:
            return None, None, None

    def search_result_check(self, soup): # returns False if no results found, True if results were found, otherwise None
        try:
            # Define the target text
            target_text = "About 0 results"

            # Find the div element with class "LHJvCe" and id "result-stats"
            result_stats_div = soup.find('div', {'class': 'LHJvCe', 'id': 'result-stats'})

            # Check if the target text is present in the div
            if result_stats_div and target_text in result_stats_div.get_text():
                return False
            else:
                return True
        except Exception as e:
            return None

    def specific_product_page(self, soup, product_code): # Returns the link to the product page ro None
        try:
            # Find the <em> tag
            em_tag = soup.find('em')
            # Check if the text inside the <em> tag is equal to product_code
            if em_tag and int(em_tag.text.strip()) == int(product_code):
                # Find the <a> tag and get the href value
                a_tag = soup.find('a', {'jsname': 'UWckNb'})
                if a_tag:
                    link = a_tag['href']
                    if "bauhaus" in link:
                        return link
        except Exception as e:
            return None
        
    def extract_product_information(self, soup): # Returns price and product name or None
        try:
            # Find the <span> element with class "price"
            price_span = soup.find('span', class_='price')
            # Extract the text content from the price_span
            price_text = price_span.get_text(strip=True) 
            price = float(price_text.replace('\xa0', ''))/100

            # Find the <span> element with class "base" and itemprop "name"
            name_span = soup.find('h1', class_='page-title').find('span', class_='base')
            # Extract the text content from the name_span
            product_name = name_span.get_text(strip=True) 

            return price, product_name
        except Exception as e:
            return None, None
            
def format_product_link(url, product_name, sheet, given_row, given_column):
    try:
        cell = sheet.cell(row=given_row, column=given_column)
        cell.value = product_name
        cell.hyperlink = url
        return None
    except Exception as e:
        return None

def format_EAN(product_code):
    try:
        if product_code is str:
                product_code = product_code.replace("'","")
                product_code = int(product_code)
        return product_code
    except Exception as e:
        return None
    
class ExcelHandler:
    def handle(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            frozen_pane = sheet.sheet_view.pane
            frozen_pane_height = int(frozen_pane.ySplit) if frozen_pane.ySplit is not None else 0
            target_column_index = 1
            blank_cell_encountered = False

            with concurrent.futures.ThreadPoolExecutor() as executor:
                # Create instances of scraper classes directly
                prisma_scraper = PrismaScraper()
                k_rauta_scraper = KRautaScraper()
                puuilo_scraper = PuuiloScraper()
                tokmanni_scraper = TokmanniScraper()
                bauhaus_scraper = BauhausScraper()
                # List of functions to execute in parallel
                functions_to_execute = [
                    prisma_scraper.search_product,
                    k_rauta_scraper.search_product,
                    puuilo_scraper.search_product,
                    tokmanni_scraper.search_product,
                    bauhaus_scraper.search_product,
                ]
                
                for row in sheet.iter_rows(min_row=frozen_pane_height + 1, max_row=sheet.max_row, min_col=1, max_col=12):
                    if target_column_index >= len(row):
                        new_column = sheet.max_column + 1
                        sheet.cell(row=1, column=new_column, value=f'Column{new_column}')

                    if row[0].value is not None:
                        # Execute functions in parallel
                        #|        PRISMA     |      K-RAUTA     |       PUUILO     |      TOKMANNI    |      BAUHAUS     |
                        #|   [0]   [1]   [2] | [0]  [1]    [2]  | [0]  [1]    [2]  | [0]  [1]   [2]   | [0]  [1]    [2]  |
                        #|PRODUCT|LINK |BRAND|LINK|PRICE|PRODUCT|LINK|PRICE|PRODUCT|LINK|PRICE|PRODUCT|LINK|PRICE|PRODUCT|
                        product_code = format_EAN(row[0].value)
                        results = list(executor.map(lambda f: f(product_code), functions_to_execute))
                        # Process results and update the sheet accordingly
                        # My first time using lambda or concurrent.futures so allow it
                        if results[0]: #PRISMA INFO TO EXCEL
                            sheet.cell(row=row[0].row, column=target_column_index + 1, value=results[0][0].upper()) # product name
                            sheet.cell(row=row[0].row, column=target_column_index + 2, value=results[0][2])         # brand
                            current_row = row[0].row
                            current_column = target_column_index + 1
                            format_product_link(results[0][1], results[0][0], sheet, current_row, current_column)   # link with product name as label
                        if results[1]: #K-RAUTA INFO TO EXCEL
                            sheet.cell(row=row[0].row, column=target_column_index + 3, value=results[1][1])         # price
                            current_row = row[0].row
                            current_column = target_column_index + 4
                            format_product_link(results[1][0], results[1][2], sheet, current_row, current_column)   # link with product name as label
                        else:
                            sheet.cell(row=row[0].row, column=target_column_index + 3, value="")                    # leave K-rauta price empty
                            sheet.cell(row=row[0].row, column=target_column_index + 4, value="")                    # leave K-rauta link empty
                        if results[4]: #BAUHAUS INFO TO EXCEL
                            sheet.cell(row=row[0].row, column=target_column_index + 5, value=results[4][1])
                            current_row = row[0].row
                            current_column = target_column_index + 6
                            format_product_link(results[4][0], results[4][2], sheet, current_row, current_column)
                        else:
                            sheet.cell(row=row[0].row, column=target_column_index + 5, value="")
                            sheet.cell(row=row[0].row, column=target_column_index + 6, value="")
                        if results[2]: #PUUILO INFO TO EXCEL
                            sheet.cell(row=row[0].row, column=target_column_index + 7, value=results[2][1])
                            current_row = row[0].row
                            current_column = target_column_index + 8
                            format_product_link(results[2][0], results[2][2], sheet, current_row, current_column)
                        else:
                            sheet.cell(row=row[0].row, column=target_column_index + 7, value="")
                            sheet.cell(row=row[0].row, column=target_column_index + 8, value="")

                        if results[3]: #TOKMANNI INFO TO EXCEL
                            sheet.cell(row=row[0].row, column=target_column_index + 9, value=results[3][1])
                            current_row = row[0].row
                            current_column = target_column_index + 10
                            format_product_link(results[3][0], results[3][2], sheet, current_row, current_column)
                        else:
                            sheet.cell(row=row[0].row, column=target_column_index + 9, value="")
                            sheet.cell(row=row[0].row, column=target_column_index + 10, value="")
                    else:
                        blank_cell_encountered = True

                    if blank_cell_encountered:
                        break

            wb.save(file_path)
            return 0
        except Exception as e:
            return 0


# Main
##################################################################################################
# JOS SÄ ET OSAA OHJELMOIDA NIIN ÄLÄ KOSKE MIHINKÄÄN PAITSI TÄHÄN JOS ON IHAN PAKKO              #
# LAITA SUN SEM EXCEL-TIEDOSTON SIJAINTI MIKÄ SISÄLTÄÄ EAN-KOODIT TON ALLA OLEVAN                #
# "file_path =" perään HEITTOMERKKIEN SISÄÄN esimerkiksi JOS SUN TIEDOSOTO ON                    #
# SUN TIETOKONEELLA SIJAINNISSA "C:/Users/Karoliina/kilpailijahintoja.xlsx" NIIN                 #
# MUUTA TOI RIVI NIIN ETTÄ SIINÄ LUKEE: file_path = "C:/Users/Karoliina/kilpailijahintoja.xlsx"  #
##################################################################################################
file_path = "C:/Users/Karoliina/kilpailijahintoja.xlsx"

start_time = time.time()
handler = ExcelHandler()
handler.handle(file_path)
end_time = time.time()
elapsed_time = end_time - start_time
# Play a sound when done
if __name__ == "__main__":
    print(f"This program took: {elapsed_time:.2f} seconds to run. Kiitos ja anteeksi.")
    for _ in range(3):
        winsound.Beep(1000, 500)
